#!/usr/bin/env python3
"""
export_sign_prior.py — export sign-prior matrices and evidence to Excel.

Output: results/sign_prior_export.xlsx
Sheets:
  1. README          — parameter choices, method notes
  2. Hamilton_prior  — symmetric sign prior (10×10) for Hamilton model
  3. gLV_prior       — directed sign prior (10×10) for gLV model
  4. Hamilton_netflow — raw numeric net-flow (symmetric)
  5. gLV_netflow      — raw numeric net-flow (directed)
  6. Evidence         — per-pair evidence breakdown (metabolites, weights)
  7. Competition      — which metabolites drive each competition signal
"""
import sys
from pathlib import Path
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_00

sys.path.insert(0, str(Path(__file__).parent))
from guild_replicator_dieckow import GUILD_ORDER, N_G
from build_net_flow_expanded import (
    net_flow_hamilton, net_flow_glv, build_net_flow_expanded,
    _guild_of, _COMPETITION_EXCLUDE, _load_suppfile,
)

OUT = Path(__file__).parent / 'results' / 'sign_prior_export.xlsx'
OUT.parent.mkdir(exist_ok=True)

GS = [g for g in GUILD_ORDER]          # full names for rows/cols
GS_SHORT = {
    'Actinobacteria':      'Actin.',
    'Bacilli':             'Bacil.',
    'Bacteroidia':         'Bact.',
    'Betaproteobacteria':  'β-Prot.',
    'Clostridia':          'Clost.',
    'Coriobacteriia':      'Corio.',
    'Fusobacteriia':       'Fusob.',
    'Gammaproteobacteria': 'γ-Prot.',
    'Negativicutes':       'Negat.',
    'Other':               'Other',
}

# ── colour palette ────────────────────────────────────────────────────────────
C_POS    = 'C6EFCE'   # green  — mutualism / positive prior
C_NEG    = 'FFC7CE'   # red    — competition / negative prior
C_ZERO   = 'FFFFFF'   # white  — unconstrained
C_DIAG   = 'D9D9D9'   # grey   — diagonal (self)
C_HEADER = '2E4057'   # dark blue header
C_HEAD_F = 'FFFFFF'   # white text on header

thin = Side(style='thin', color='999999')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def hdr_font():
    return Font(bold=True, color=C_HEAD_F)

def bold():
    return Font(bold=True)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)


# ── build data ────────────────────────────────────────────────────────────────
h_net  = net_flow_hamilton(use_agora=False, competition_weight=0.5)
g_net  = net_flow_glv(use_agora=False, competition_weight=0.5)
h_sign = np.sign(h_net).astype(int)
g_sign = np.sign(g_net).astype(int)

# per-pair evidence (cross-feeding + competition separately)
df_sf = _load_suppfile()
gi = {g: idx for idx, g in enumerate(GUILD_ORDER)}
cons_rel = {'USES', 'DEPENDS_ON', 'HYDROLYSES', 'DEGRADES'}
prod_rel = {'PRODUCES', 'RELEASES'}

def szaf_weight(row):
    is_exp = str(row.get('EVIDENCE', '')).strip() == 'experimental'
    kegg   = str(row.get('KEGG', ''))
    hmdb   = str(row.get('HMDB_ID', ''))
    has_db = kegg not in ('n/a', '', 'nan', 'NaN') or 'HMDB' in hmdb
    return (2.0 if has_db else 1.5) if is_exp else 1.0

evidence_rows = []      # for Evidence sheet
competition_rows = []   # for Competition sheet

for met in df_sf['OBJECT'].unique():
    mdf = df_sf[df_sf['OBJECT'] == met]
    w   = float(mdf.apply(szaf_weight, axis=1).max())
    obj_type = str(mdf['OBJECT_TYPE'].iloc[0])
    evidence_type = 'L1-exp' if mdf.apply(
        lambda r: str(r.get('EVIDENCE','')).strip()=='experimental', axis=1).any() else 'L2-pred'

    prod, cons = set(), set()
    for _, row in mdf.iterrows():
        g   = _guild_of(row['TAXON'])
        rel = str(row['RELATIONSHIP'])
        if g is None or g not in gi:
            continue
        if rel in prod_rel:
            prod.add(g)
        elif rel in cons_rel:
            cons.add(g)

    # cross-feeding evidence
    for src in prod:
        for tgt in cons:
            if src == tgt:
                continue
            evidence_rows.append({
                'guild_i': tgt, 'guild_j': src,
                'metabolite': met, 'object_type': obj_type,
                'relationship': 'cross-feeding (j→i)',
                'layer': evidence_type, 'weight': w,
                'sign': '+1',
            })

    # competition evidence
    if met not in _COMPETITION_EXCLUDE:
        cons_list = sorted(cons)
        for ii, ga in enumerate(cons_list):
            for gb in cons_list[ii+1:]:
                competition_rows.append({
                    'guild_a': ga, 'guild_b': gb,
                    'metabolite': met, 'object_type': obj_type,
                    'layer': evidence_type, 'weight_per_guild': w * 0.5,
                    'excluded': 'no',
                })
    elif met in _COMPETITION_EXCLUDE:
        cons_list = sorted(cons)
        for ii, ga in enumerate(cons_list):
            for gb in cons_list[ii+1:]:
                competition_rows.append({
                    'guild_a': ga, 'guild_b': gb,
                    'metabolite': met, 'object_type': obj_type,
                    'layer': evidence_type, 'weight_per_guild': 0.0,
                    'excluded': f'YES — environmental variable',
                })

df_ev   = pd.DataFrame(evidence_rows)
df_comp = pd.DataFrame(competition_rows)


# ── helpers ───────────────────────────────────────────────────────────────────
def write_matrix_sheet(ws, matrix, title, mode='sign'):
    """Write a 10×10 guild matrix with colour coding."""
    ws.title = title

    # title row
    ws.merge_cells('A1:K1')
    c = ws['A1']
    c.value = title
    c.font  = Font(bold=True, size=13, color=C_HEAD_F)
    c.fill  = fill(C_HEADER)
    c.alignment = center()
    ws.row_dimensions[1].height = 22

    # header row (col labels)
    ws.cell(2, 1).value = 'row i  →  col j'
    ws.cell(2, 1).font  = bold()
    ws.cell(2, 1).fill  = fill(C_HEADER)
    ws.cell(2, 1).font  = Font(bold=True, color=C_HEAD_F)
    ws.cell(2, 1).alignment = center()
    for j, g in enumerate(GS):
        c = ws.cell(2, j + 2)
        c.value = GS_SHORT[g]
        c.font  = Font(bold=True, color=C_HEAD_F)
        c.fill  = fill(C_HEADER)
        c.alignment = center()
        ws.column_dimensions[get_column_letter(j + 2)].width = 8

    ws.column_dimensions['A'].width = 16

    # data rows
    for i, gi_name in enumerate(GS):
        r = i + 3
        c = ws.cell(r, 1)
        c.value = GS_SHORT[gi_name]
        c.font  = Font(bold=True, color=C_HEAD_F)
        c.fill  = fill(C_HEADER)
        c.alignment = center()
        ws.row_dimensions[r].height = 18

        for j in range(N_G):
            v = float(matrix[i, j])
            cell = ws.cell(r, j + 2)
            if mode == 'sign':
                cell.value = int(v)
                if i == j:
                    cell.fill = fill(C_DIAG)
                elif v > 0:
                    cell.fill = fill(C_POS)
                elif v < 0:
                    cell.fill = fill(C_NEG)
                else:
                    cell.fill = fill(C_ZERO)
                cell.alignment = center()
                cell.font = Font(bold=(v != 0 and i != j))
            else:  # numeric
                cell.value = round(v, 2)
                if i == j:
                    cell.fill = fill(C_DIAG)
                elif v > 0:
                    cell.fill = fill(C_POS)
                elif v < 0:
                    cell.fill = fill(C_NEG)
                else:
                    cell.fill = fill(C_ZERO)
                cell.alignment = center()
                cell.number_format = '0.00'
            cell.border = border

    # legend
    leg_r = N_G + 4
    ws.cell(leg_r, 1).value = 'Legend:'
    ws.cell(leg_r, 1).font  = bold()
    items = [('+1 / positive', C_POS, 'mutualism / cross-feeding dominates'),
             ('-1 / negative', C_NEG, 'competition dominates'),
             ('0',             C_ZERO, 'unconstrained (no evidence or contradictory)'),
             ('diagonal',      C_DIAG, 'self-interaction (not used in prior)')]
    for k, (label, col, desc) in enumerate(items):
        ws.cell(leg_r + k + 1, 1).value = label
        ws.cell(leg_r + k + 1, 1).fill  = fill(col)
        ws.cell(leg_r + k + 1, 1).border = border
        ws.cell(leg_r + k + 1, 2).value = desc
        ws.merge_cells(start_row=leg_r+k+1, start_column=2,
                       end_row=leg_r+k+1, end_column=6)


def write_readme(ws):
    ws.title = 'README'
    rows = [
        ('Sign-prior export for Hamilton and gLV models', True, 14),
        ('', False, 11),
        ('METHOD', True, 12),
        ('Evidence is aggregated from three layers:', False, 11),
        ('  L1 experimental + KEGG/HMDB annotation  →  weight 2.0', False, 11),
        ('  L1 experimental, no DB annotation        →  weight 1.5', False, 11),
        ('  L2 predicted                             →  weight 1.0', False, 11),
        ('  L3 AGORA2 FBA cross-feeding              →  weight 0.5  (not included in this export)', False, 11),
        ('', False, 11),
        ('Cross-feeding signal:', True, 11),
        ('  If guild j PRODUCES metabolite M, and guild i USES M  →  pos[i,j] += w', False, 11),
        ('  sign_prior[i,j] = sign(pos[i,j] − neg[i,j])', False, 11),
        ('', False, 11),
        ('Competition signal (alpha = 0.5):', True, 11),
        ('  If guild i and guild j both USE metabolite M  →  neg[i,j] += w × 0.5,  neg[j,i] += w × 0.5', False, 11),
        ('  Environmental variables excluded: O₂, CO₂, H₂O₂', False, 11),
        ('', False, 11),
        ('IS_INHIBITED_BY note:', True, 11),
        ('  22 rows in Szafranski file; 20 are "IS_INHIBITED_BY oxygen".', False, 11),
        ('  No guild PRODUCES oxygen → these 20 rows contribute zero neg signal.', False, 11),
        ('  Only H₂O₂ fires (2 pairs: Bacilli inhibits Bacteroidia and Actinobacteria).', False, 11),
        ('', False, 11),
        ('MODEL-SPECIFIC VERSIONS', True, 12),
        ('Hamilton model: A is symmetric (A[i,j] = A[j,i])', True, 11),
        ('  → net_flow is symmetrized: net_sym = (net + net.T) / 2', False, 11),
        ('  → Amensalism (+/−) signals in opposite directions cancel to 0 (unconstrained)', False, 11),
        ('  → Sheet: Hamilton_prior, Hamilton_netflow', False, 11),
        ('', False, 11),
        ('gLV model: A is a full N×N matrix (no symmetry constraint)', True, 11),
        ('  → net_flow is directed: A[i,j] and A[j,i] constrained independently', False, 11),
        ('  → Amensalism (+/−) is preserved as a meaningful biological signal', False, 11),
        ('  → Sheet: gLV_prior, gLV_netflow', False, 11),
        ('', False, 11),
        ('CODE', True, 12),
        ('  build_net_flow_expanded.py — core function', False, 11),
        ('  net_flow_hamilton(use_agora, competition_weight) → symmetric matrix', False, 11),
        ('  net_flow_glv(use_agora, competition_weight)      → directed matrix', False, 11),
        ('', False, 11),
        ('PARAMETERS USED IN THIS EXPORT', True, 12),
        ('  competition_weight (alpha) = 0.5', False, 11),
        ('  use_agora                  = False  (AGORA FBA not included)', False, 11),
        ('  symmetrize (Hamilton)      = True', False, 11),
        ('  symmetrize (gLV)           = False', False, 11),
    ]
    ws.column_dimensions['A'].width = 90
    for r, (text, is_bold, size) in enumerate(rows, 1):
        c = ws.cell(r, 1)
        c.value = text
        c.font  = Font(bold=is_bold, size=size)
        c.alignment = Alignment(wrap_text=False)


# ── build workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# Sheet 1: README
ws_readme = wb.active
write_readme(ws_readme)

# Sheet 2: Hamilton sign prior
ws_h = wb.create_sheet()
write_matrix_sheet(ws_h, h_sign, 'Hamilton_prior', mode='sign')

# Sheet 3: gLV sign prior (show full N×N directed)
ws_g = wb.create_sheet()
write_matrix_sheet(ws_g, g_sign, 'gLV_prior', mode='sign')

# Sheet 4: Hamilton net-flow (numeric)
ws_hn = wb.create_sheet()
write_matrix_sheet(ws_hn, h_net, 'Hamilton_netflow', mode='numeric')

# Sheet 5: gLV net-flow (numeric)
ws_gn = wb.create_sheet()
write_matrix_sheet(ws_gn, g_net, 'gLV_netflow', mode='numeric')

# Sheet 6: Evidence breakdown
ws_ev = wb.create_sheet('Evidence')
if not df_ev.empty:
    cols = ['guild_i', 'guild_j', 'metabolite', 'object_type',
            'relationship', 'layer', 'weight', 'sign']
    df_ev_out = df_ev[cols].sort_values(['guild_i', 'guild_j', 'metabolite'])
    # header
    for ci, col in enumerate(cols, 1):
        c = ws_ev.cell(1, ci)
        c.value = col
        c.font  = Font(bold=True, color=C_HEAD_F)
        c.fill  = fill(C_HEADER)
        c.alignment = center()
    for ri, row in enumerate(df_ev_out.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            ws_ev.cell(ri, ci).value = val
    for ci, col in enumerate(cols, 1):
        ws_ev.column_dimensions[get_column_letter(ci)].width = max(
            len(col) + 2,
            df_ev_out.iloc[:, ci-1].astype(str).str.len().max() + 2
        )
    ws_ev.auto_filter.ref = f'A1:{get_column_letter(len(cols))}1'

# Sheet 7: Competition pairs
ws_cp = wb.create_sheet('Competition')
if not df_comp.empty:
    cols_c = ['guild_a', 'guild_b', 'metabolite', 'object_type',
              'layer', 'weight_per_guild', 'excluded']
    df_cp_out = df_comp[cols_c].sort_values(['guild_a', 'guild_b'])
    for ci, col in enumerate(cols_c, 1):
        c = ws_cp.cell(1, ci)
        c.value = col
        c.font  = Font(bold=True, color=C_HEAD_F)
        c.fill  = fill(C_HEADER)
        c.alignment = center()
    for ri, row in enumerate(df_cp_out.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            cell = ws_cp.cell(ri, ci)
            cell.value = val
            if str(val).startswith('YES'):
                cell.fill = fill('FFF2CC')   # yellow — excluded
    for ci, col in enumerate(cols_c, 1):
        ws_cp.column_dimensions[get_column_letter(ci)].width = max(
            len(col) + 2,
            df_cp_out.iloc[:, ci-1].astype(str).str.len().max() + 2
        )
    ws_cp.auto_filter.ref = f'A1:{get_column_letter(len(cols_c))}1'

wb.save(OUT)
print(f'Saved: {OUT}')
n_pos_h = sum(1 for i in range(N_G) for j in range(i+1,N_G) if h_sign[i,j]>0)
n_neg_h = sum(1 for i in range(N_G) for j in range(i+1,N_G) if h_sign[i,j]<0)
print(f'  Hamilton: {n_pos_h} mutualism, {n_neg_h} competition, '
      f'{45-n_pos_h-n_neg_h} unconstrained  (of 45 pairs)')
n_pos_g = int((g_sign>0).sum()) - int(np.diag(g_sign>0).sum())
n_neg_g = int((g_sign<0).sum())
print(f'  gLV:      {n_pos_g} positive, {n_neg_g} negative, '
      f'{N_G*(N_G-1)-n_pos_g-n_neg_g} unconstrained  (of {N_G*(N_G-1)} directed)')
print(f'  Evidence rows: {len(df_ev)}')
print(f'  Competition rows: {len(df_comp)}  (incl. excluded)')
