#!/usr/bin/env python3
"""
build_neo4j_import.py — Convert Dieckow SI + Szafranski MOESM3 to Neo4j import CSVs.

Graph schema:
  (:Species  {id, name, HMT, eHOMD, class})
  (:Substrate {id, name, sub_type, HMDB_ID, HMDB_url, EC, KEGG})
      sub_type ∈ {metabolite, enzyme, biofilm_component, taxon, cell_component}

  (Species)-[:PRODUCES|USES|IS_INHIBITED_BY|DEGRADES|
              DEPENDS_ON|HYDROLYSES|RELEASES|IS_A_HOST_FOR
              {evidence, figure, pubmed1, pubmed2, references, comments}]->(Substrate)

  (Species)-[:SUPPORTS {oral, strain_details, co_culture, reference}]->(Species)  ← MOESM3

Outputs (neo4j-admin import format):
  results/neo4j_import/nodes_species.csv
  results/neo4j_import/nodes_substrate.csv
  results/neo4j_import/rels_species_substrate.csv
  results/neo4j_import/rels_species_supports.csv
  results/neo4j_import/import.cypher          ← MERGE-based, usable in Browser

Usage:
  python build_neo4j_import.py
  # then neo4j-admin database import full --overwrite-destination \\
  #   --nodes=Species=results/neo4j_import/nodes_species.csv \\
  #   --nodes=Substrate=results/neo4j_import/nodes_substrate.csv \\
  #   --relationships=results/neo4j_import/rels_species_substrate.csv \\
  #   --relationships=results/neo4j_import/rels_species_supports.csv \\
  #   oral_biofilm
"""
import csv
import re
import openpyxl
from pathlib import Path

DIECKOW_XLSX = Path('/home/nishioka/IKM_Hiwi/nife/Datasets/20260416_AbutmentPapernpjBiofilmsDieckow_SI_Relationships.xlsx')
MOESM3_XLSX  = Path('/home/nishioka/IKM_Hiwi/nife/Szafranski_Published_Work/Szafranski_Published_Work/public_data/Springer/10.1038_s41467-025-66804-7/MOESM3_ESM.xlsx')
OUT_DIR      = Path('/home/nishioka/IKM_Hiwi/nife/results/neo4j_import')


def slugify(s):
    return re.sub(r'[^a-z0-9]+', '_', str(s).lower()).strip('_')


def load_dieckow():
    wb = openpyxl.load_workbook(DIECKOW_XLSX, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    idx = {h: i for i, h in enumerate(header)}
    records = []
    for r in rows[1:]:
        if not r[idx['TAXON']]:
            continue
        records.append({
            'taxon':     str(r[idx['TAXON']]).strip(),
            'HMT':       str(r[idx['HMT']]).strip() if r[idx['HMT']] else '',
            'eHOMD':     str(r[idx['eHOMD']]).strip() if r[idx['eHOMD']] else '',
            'rel':       str(r[idx['RELATIONSHIP']]).strip(),
            'object':    str(r[idx['OBJECT']]).strip(),
            'obj_type':  str(r[idx['OBJECT_TYPE']]).strip() if r[idx['OBJECT_TYPE']] else 'metabolite',
            'HMDB_ID':   str(r[idx['HMDB_ID']]).strip() if r[idx['HMDB_ID']] else '',
            'HMDB_url':  str(r[idx['HMDB']]).strip() if r[idx['HMDB']] else '',
            'EC':        str(r[idx['EC']]).strip() if r[idx['EC']] else '',
            'KEGG':      str(r[idx['KEGG']]).strip() if r[idx['KEGG']] else '',
            'evidence':  str(r[idx['EVIDENCE']]).strip() if r[idx['EVIDENCE']] else '',
            'comments':  str(r[idx['CURATORS_COMMENTS']]).strip() if r[idx['CURATORS_COMMENTS']] else '',
            'references':str(r[idx['REFERENCES']]).strip() if r[idx['REFERENCES']] else '',
            'pubmed1':   str(r[idx['PubMed1']]).strip() if r[idx['PubMed1']] else '',
            'pubmed2':   str(r[idx['PubMed2']]).strip() if r[idx['PubMed2']] else '',
            'figure':    str(r[idx['Figure']]).strip() if r[idx['Figure']] else '',
        })
    return records


def load_moesm3():
    wb = openpyxl.load_workbook(MOESM3_XLSX, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    idx = {h: i for i, h in enumerate(header)}
    records = []
    for r in rows[1:]:
        if not r[0]:
            continue
        records.append({
            'helper':       str(r[idx['Helper species']]).strip(),
            'helper_HMT':   str(r[idx['HMT no. of helper']]) if r[idx['HMT no. of helper']] else '',
            'helper_eHOMD': str(r[idx['eHOMD link for helper']]) if r[idx['eHOMD link for helper']] else '',
            'helper_class': str(r[idx['Class of helper']]) if r[idx['Class of helper']] else '',
            'supported':    str(r[idx['Supported species']]).strip(),
            'sup_HMT':      str(r[idx['HMT no. of supported species']]) if r[idx['HMT no. of supported species']] else '',
            'sup_eHOMD':    str(r[idx['eHOMD link for supported species']]) if r[idx['eHOMD link for supported species']] else '',
            'sup_class':    str(r[idx['Class of supported species']]) if r[idx['Class of supported species']] else '',
            'oral':         str(r[idx['Related to oral microbiome']]) if r[idx['Related to oral microbiome']] else '',
            'strain':       str(r[idx['Strain details']]) if r[idx['Strain details']] else '',
            'cocult':       str(r[idx['Co-culture conditions']]) if r[idx['Co-culture conditions']] else '',
            'comments':     str(r[idx['Comments']]) if r[idx['Comments']] else '',
            'reference':    str(r[idx['Reference (DOI)']]) if r[idx['Reference (DOI)']] else '',
        })
    return records


def build_graph(dieckow, moesm3):
    species   = {}   # slug → {id, name, HMT, eHOMD, class}
    substrates = {}  # slug → {id, name, sub_type, HMDB_ID, HMDB_url, EC, KEGG}
    ss_rels   = []   # species → substrate
    sp_rels   = []   # species → species (SUPPORTS)

    def add_species(name, HMT='', eHOMD='', cls=''):
        sid = 'sp_' + slugify(name)
        if sid not in species:
            species[sid] = {'id': sid, 'name': name, 'HMT': HMT,
                            'eHOMD': eHOMD, 'class': cls}
        return sid

    def add_substrate(name, sub_type, HMDB_ID='', HMDB_url='', EC='', KEGG=''):
        bid = 'sub_' + slugify(name)
        if bid not in substrates:
            substrates[bid] = {'id': bid, 'name': name, 'sub_type': sub_type,
                               'HMDB_ID': HMDB_ID, 'HMDB_url': HMDB_url,
                               'EC': EC, 'KEGG': KEGG}
        return bid

    # Dieckow rows
    for r in dieckow:
        sid = add_species(r['taxon'], r['HMT'], r['eHOMD'])
        bid = add_substrate(r['object'], r['obj_type'].replace(' ', '_'),
                            r['HMDB_ID'], r['HMDB_url'], r['EC'], r['KEGG'])
        ss_rels.append({
            ':START_ID': sid,
            ':END_ID':   bid,
            ':TYPE':     r['rel'],
            'evidence':  r['evidence'],
            'figure':    r['figure'],
            'pubmed1':   r['pubmed1'],
            'pubmed2':   r['pubmed2'],
            'references':r['references'],
            'comments':  r['comments'],
        })

    # MOESM3 SUPPORTS edges
    for r in moesm3:
        h_sid = add_species(r['helper'],    r['helper_HMT'], r['helper_eHOMD'], r['helper_class'])
        s_sid = add_species(r['supported'], r['sup_HMT'],    r['sup_eHOMD'],    r['sup_class'])
        sp_rels.append({
            ':START_ID':    h_sid,
            ':END_ID':      s_sid,
            ':TYPE':        'SUPPORTS',
            'oral':         r['oral'],
            'strain_details': r['strain'],
            'co_culture':   r['cocult'],
            'comments':     r['comments'],
            'reference':    r['reference'],
        })

    return species, substrates, ss_rels, sp_rels


def write_csvs(species, substrates, ss_rels, sp_rels):
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    def wcsv(path, header, rows):
        with open(path, 'w', newline='') as f:
            w = csv.writer(f)
            w.writerow(header)
            w.writerows(rows)

    wcsv(OUT_DIR / 'nodes_species.csv',
         [':ID', 'name:STRING', 'HMT:STRING', 'eHOMD:STRING', 'class:STRING', ':LABEL'],
         [[v['id'], v['name'], v['HMT'], v['eHOMD'], v['class'], 'Species']
          for v in species.values()])

    wcsv(OUT_DIR / 'nodes_substrate.csv',
         [':ID', 'name:STRING', 'sub_type:STRING', 'HMDB_ID:STRING',
          'HMDB_url:STRING', 'EC:STRING', 'KEGG:STRING', ':LABEL'],
         [[v['id'], v['name'], v['sub_type'], v['HMDB_ID'],
           v['HMDB_url'], v['EC'], v['KEGG'], 'Substrate']
          for v in substrates.values()])

    wcsv(OUT_DIR / 'rels_species_substrate.csv',
         [':START_ID', ':END_ID', ':TYPE', 'evidence:STRING', 'figure:STRING',
          'pubmed1:STRING', 'pubmed2:STRING', 'references:STRING', 'comments:STRING'],
         [[r[':START_ID'], r[':END_ID'], r[':TYPE'], r['evidence'], r['figure'],
           r['pubmed1'], r['pubmed2'], r['references'], r['comments']]
          for r in ss_rels])

    wcsv(OUT_DIR / 'rels_species_supports.csv',
         [':START_ID', ':END_ID', ':TYPE', 'oral:STRING', 'strain_details:STRING',
          'co_culture:STRING', 'comments:STRING', 'reference:STRING'],
         [[r[':START_ID'], r[':END_ID'], r[':TYPE'], r['oral'], r['strain_details'],
           r['co_culture'], r['comments'], r['reference']]
          for r in sp_rels])

    print(f"  nodes_species.csv      : {len(species)} nodes")
    print(f"  nodes_substrate.csv    : {len(substrates)} nodes")
    print(f"  rels_species_substrate : {len(ss_rels)} edges")
    print(f"  rels_species_supports  : {len(sp_rels)} edges")


def write_cypher(species, substrates, ss_rels, sp_rels):
    """Generate MERGE-based Cypher for Neo4j Browser import."""
    lines = [
        "// === Oral Biofilm Knowledge Graph ===",
        "// Source: Dieckow 2025 SI + Szafranski 2025 Nat Comms MOESM3",
        "// Import via :play or paste in Neo4j Browser",
        "",
        "// Constraints (run once)",
        "CREATE CONSTRAINT species_id IF NOT EXISTS FOR (s:Species) REQUIRE s.id IS UNIQUE;",
        "CREATE CONSTRAINT substrate_id IF NOT EXISTS FOR (s:Substrate) REQUIRE s.id IS UNIQUE;",
        "",
        "// Species nodes",
    ]
    for v in species.values():
        name  = v['name'].replace("'", "\\'")
        ehomd = v['eHOMD'].replace("'", "\\'")
        lines.append(
            f"MERGE (:Species {{id:'{v['id']}', name:'{name}', "
            f"HMT:'{v['HMT']}', eHOMD:'{ehomd}', class:'{v['class']}'}});"
        )

    lines += ["", "// Substrate nodes"]
    for v in substrates.values():
        name = v['name'].replace("'", "\\'")
        lines.append(
            f"MERGE (:Substrate {{id:'{v['id']}', name:'{name}', "
            f"sub_type:'{v['sub_type']}', HMDB_ID:'{v['HMDB_ID']}', "
            f"EC:'{v['EC']}', KEGG:'{v['KEGG']}'}});"
        )

    lines += ["", "// Species→Substrate relationships"]
    for r in ss_rels:
        ev  = r['evidence'].replace("'", "\\'")
        fig = r['figure'].replace("'", "\\'")
        p1  = r['pubmed1'].replace("'", "\\'")
        lines.append(
            f"MATCH (a:Species {{id:'{r[':START_ID']}'}}), "
            f"(b:Substrate {{id:'{r[':END_ID']}'}}) "
            f"MERGE (a)-[:{r[':TYPE']} {{evidence:'{ev}', figure:'{fig}', pubmed1:'{p1}'}}]->(b);"
        )

    lines += ["", "// Species→Species SUPPORTS relationships (MOESM3)"]
    for r in sp_rels:
        oral = r['oral'].replace("'", "\\'")
        ref  = r['reference'].replace("'", "\\'")
        lines.append(
            f"MATCH (a:Species {{id:'{r[':START_ID']}'}}), "
            f"(b:Species {{id:'{r[':END_ID']}'}}) "
            f"MERGE (a)-[:SUPPORTS {{oral:'{oral}', reference:'{ref}'}}]->(b);"
        )

    cypher_path = OUT_DIR / 'import.cypher'
    with open(cypher_path, 'w') as f:
        f.write('\n'.join(lines) + '\n')
    print(f"  import.cypher          : {len(lines)} lines")


def write_load_csv_cypher():
    """
    Generate LOAD CSV Cypher for Neo4j Desktop.
    Files must be placed in Neo4j's import folder.
    Run each block separately in Neo4j Browser.
    """
    lines = [
        "// ============================================================",
        "// Oral Biofilm Knowledge Graph — Neo4j Desktop LOAD CSV import",
        "// Source: Dieckow 2025 npj Biofilms SI + Szafranski 2025 Nat Comms MOESM3",
        "//",
        "// Step 1: Copy these 4 files to your Neo4j import folder:",
        "//   nodes_species.csv",
        "//   nodes_substrate.csv",
        "//   rels_species_substrate.csv",
        "//   rels_species_supports.csv",
        "//",
        "// Step 2: Run each block below in Neo4j Browser (one at a time).",
        "// ============================================================",
        "",
        "// --- Block 1: Constraints ---",
        "CREATE CONSTRAINT species_id IF NOT EXISTS FOR (s:Species) REQUIRE s.id IS UNIQUE;",
        "CREATE CONSTRAINT substrate_id IF NOT EXISTS FOR (s:Substrate) REQUIRE s.id IS UNIQUE;",
        "",
        "// --- Block 2: Species nodes ---",
        "LOAD CSV WITH HEADERS FROM 'file:///nodes_species.csv' AS row",
        "MERGE (s:Species {id: row.`:ID`})",
        "SET s.name   = row.`name:STRING`,",
        "    s.HMT    = row.`HMT:STRING`,",
        "    s.eHOMD  = row.`eHOMD:STRING`,",
        "    s.class  = row.`class:STRING`;",
        "",
        "// --- Block 3: Substrate nodes ---",
        "LOAD CSV WITH HEADERS FROM 'file:///nodes_substrate.csv' AS row",
        "MERGE (s:Substrate {id: row.`:ID`})",
        "SET s.name     = row.`name:STRING`,",
        "    s.sub_type = row.`sub_type:STRING`,",
        "    s.HMDB_ID  = row.`HMDB_ID:STRING`,",
        "    s.HMDB_url = row.`HMDB_url:STRING`,",
        "    s.EC       = row.`EC:STRING`,",
        "    s.KEGG     = row.`KEGG:STRING`;",
        "",
        "// --- Block 4: Species→Substrate relationships ---",
        "LOAD CSV WITH HEADERS FROM 'file:///rels_species_substrate.csv' AS row",
        "MATCH (a:Species  {id: row.`:START_ID`})",
        "MATCH (b:Substrate{id: row.`:END_ID`})",
        "CALL apoc.merge.relationship(a, row.`:TYPE`,",
        "  {evidence: row.`evidence:STRING`,",
        "   figure:   row.`figure:STRING`,",
        "   pubmed1:  row.`pubmed1:STRING`,",
        "   pubmed2:  row.`pubmed2:STRING`,",
        "   references: row.`references:STRING`,",
        "   comments: row.`comments:STRING`},",
        "  {}, b) YIELD rel RETURN count(rel);",
        "",
        "// --- Block 4 (no APOC fallback) ---",
        "// If APOC is not installed, run this instead of Block 4:",
        "// LOAD CSV WITH HEADERS FROM 'file:///rels_species_substrate.csv' AS row",
        "// MATCH (a:Species  {id: row.`:START_ID`})",
        "// MATCH (b:Substrate{id: row.`:END_ID`})",
        "// FOREACH (_ IN CASE row.`:TYPE` WHEN 'PRODUCES'       THEN [1] ELSE [] END |",
        "//   MERGE (a)-[:PRODUCES       {evidence:row.`evidence:STRING`}]->(b))",
        "// FOREACH (_ IN CASE row.`:TYPE` WHEN 'USES'           THEN [1] ELSE [] END |",
        "//   MERGE (a)-[:USES           {evidence:row.`evidence:STRING`}]->(b))",
        "// FOREACH (_ IN CASE row.`:TYPE` WHEN 'IS_INHIBITED_BY' THEN [1] ELSE [] END |",
        "//   MERGE (a)-[:IS_INHIBITED_BY{evidence:row.`evidence:STRING`}]->(b))",
        "// FOREACH (_ IN CASE row.`:TYPE` WHEN 'DEGRADES'        THEN [1] ELSE [] END |",
        "//   MERGE (a)-[:DEGRADES       {evidence:row.`evidence:STRING`}]->(b))",
        "// FOREACH (_ IN CASE row.`:TYPE` WHEN 'DEPENDS_ON'      THEN [1] ELSE [] END |",
        "//   MERGE (a)-[:DEPENDS_ON     {evidence:row.`evidence:STRING`}]->(b))",
        "// FOREACH (_ IN CASE row.`:TYPE` WHEN 'HYDROLYSES'      THEN [1] ELSE [] END |",
        "//   MERGE (a)-[:HYDROLYSES     {evidence:row.`evidence:STRING`}]->(b))",
        "// FOREACH (_ IN CASE row.`:TYPE` WHEN 'RELEASES'        THEN [1] ELSE [] END |",
        "//   MERGE (a)-[:RELEASES       {evidence:row.`evidence:STRING`}]->(b))",
        "// FOREACH (_ IN CASE row.`:TYPE` WHEN 'IS_A_HOST_FOR'   THEN [1] ELSE [] END |",
        "//   MERGE (a)-[:IS_A_HOST_FOR  {evidence:row.`evidence:STRING`}]->(b));",
        "",
        "// --- Block 5: Species→Species SUPPORTS (MOESM3) ---",
        "LOAD CSV WITH HEADERS FROM 'file:///rels_species_supports.csv' AS row",
        "MATCH (a:Species{id: row.`:START_ID`})",
        "MATCH (b:Species{id: row.`:END_ID`})",
        "MERGE (a)-[r:SUPPORTS]->(b)",
        "SET r.oral           = row.`oral:STRING`,",
        "    r.strain_details = row.`strain_details:STRING`,",
        "    r.co_culture     = row.`co_culture:STRING`,",
        "    r.comments       = row.`comments:STRING`,",
        "    r.reference      = row.`reference:STRING`;",
        "",
        "// --- Verification queries ---",
        "MATCH (n) RETURN labels(n)[0] AS label, count(n) AS count;",
        "MATCH ()-[r]->() RETURN type(r) AS rel_type, count(r) AS count ORDER BY count DESC;",
    ]

    path = OUT_DIR / 'load_csv.cypher'
    with open(path, 'w') as f:
        f.write('\n'.join(lines) + '\n')
    print(f"  load_csv.cypher        : {len(lines)} lines  ← Neo4j Desktop 用")


def main():
    print("Loading Dieckow SI xlsx...")
    dieckow = load_dieckow()
    print(f"  {len(dieckow)} rows loaded")

    print("Loading MOESM3 helper network...")
    moesm3 = load_moesm3()
    print(f"  {len(moesm3)} rows loaded")

    print("Building graph...")
    species, substrates, ss_rels, sp_rels = build_graph(dieckow, moesm3)

    print("Writing CSVs...")
    write_csvs(species, substrates, ss_rels, sp_rels)

    print("Writing Cypher...")
    write_cypher(species, substrates, ss_rels, sp_rels)

    print("Writing LOAD CSV Cypher (Neo4j Desktop)...")
    write_load_csv_cypher()

    print(f"\nDone. Output: {OUT_DIR}/")
    print("\nnео4j-admin import command:")
    print(f"  neo4j-admin database import full --overwrite-destination \\")
    print(f"    --nodes=Species={OUT_DIR}/nodes_species.csv \\")
    print(f"    --nodes=Substrate={OUT_DIR}/nodes_substrate.csv \\")
    print(f"    --relationships={OUT_DIR}/rels_species_substrate.csv \\")
    print(f"    --relationships={OUT_DIR}/rels_species_supports.csv \\")
    print(f"    oral_biofilm")


if __name__ == '__main__':
    main()
